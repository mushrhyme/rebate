"""SAP 엑셀 내보내기 API — 확정된 문서 선택 → 합산 엑셀 다운로드."""
import io
import json
from datetime import date
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string
from pydantic import BaseModel

from ...core.auth import get_current_user
from ...core.config import get_settings
from ...db.queries import list_documents

router = APIRouter(prefix="/api/v3/sap", tags=["sap"])

# ── 컬럼 정의 (output-format.md B~BB) ────────────────────────────────────────
# (엑셀 열 문자, 한국어 헤더, 일본어 헤더, JSON 키)
# JSON 키: None = 빈 열
_COLS = [
    ("B",  "판매처 원문명",        "受注先",                     "受注先"),
    ("C",  "판매처코드",            "受注先コード",               "受注先コード"),
    ("D",  "SP 영업사원",           "担当者",                     "担当者"),
    ("E",  "판매처코드(이관포함)",  "受注先コード（移管含む）",   None),
    ("F",  "이전판매처",            "移管受注先",                 None),
    ("G",  "이관후 담당자",         "移管後担当者",               None),
    ("H",  "로트아웃 판별용",       "ロットアウト判別用",         "ロットアウト判別用"),
    ("I",  "PY",                    "PY",                         None),
    ("J",  "소매처코드",            "代表スーパー",               "代表スーパー"),
    ("K",  "소매처명",              "スーパー",                   "スーパー"),
    ("L",  "제품명",                "商品名",                     "商品名"),
    ("M",  "제품코드",              "商品コード",                 "商品コード"),
    ("N",  "케이스 입수",           "ケース (入数)",              "ケース入数"),
    ("O",  "볼 입수",               "ボール (入数)",              "ボール入数"),
    ("P",  "케이스 수량",           "ケース (数量)",              "ケース"),
    ("Q",  "ケース計",              "*",                          "Q"),
    ("R",  "볼 수량",               "ボール (数量)",              "ボール"),
    ("S",  "ボール計",              "*",                          "S"),
    ("T",  "낱개 수량",             "バラ",                       "バラ"),
    ("U",  "총 개수",               "個数計",                     "個数計"),
    ("V",  "케이스 환산",           "ケース計",                   "ケース計"),
    ("W",  "발생월",                "発生月",                     "発生月"),
    ("X",  "시키리(단가)",          "仕切",                       "仕切"),
    ("Y",  "시키리 매출",           "仕切売上",                   "仕切売上"),
    ("Z",  "조건1(팩)",             "条件1（パック）",            "条件1（パック）"),
    ("AA", "조건2(팩)",             "条件2（パック）",            "条件2（パック）"),
    ("AB", "조건1(볼)",             "条件1（ボール）",            "条件1（ボール）"),
    ("AC", "조건2(볼)",             "条件2（ボール）",            "条件2（ボール）"),
    ("AD", "조건1(케이스)",         "条件1（ケース）",            "条件1（ケース）"),
    ("AE", "조건2(케이스)",         "条件2（ケース）",            "条件2（ケース）"),
    ("AF", "팩조건1(단위통합)",     "パック条件1",                "AF"),
    ("AG", "팩조건2(단위통합)",     "パック条件2",                "AG"),
    ("AH", "NET",                   "NET",                        "NET"),
    ("AI", "팩조건1 합계",          "パック条件1合計",            None),
    ("AJ", "팩조건2 합계",          "パック条件2合計",            None),
    ("AK", "차액",                  "差額",                       None),
    ("AL", "지급해야 될 금액",      "未収金額合計",               "未収金額合計"),
    ("AM", "상정 납가",             "想定納価",                   None),
    ("AN", "피%(×납가)①",         "ﾌｨｰ％(×納価）①",          None),
    ("AO", "피%(×납가)②",         "ﾌｨｰ％(×納価）②",          None),
    ("AP", "최종 NET",              "最終NET",                    None),
    ("AQ", "최종 NET 매출",         "最終NET売上",                None),
    ("AR", "대진료",                "大陳料",                     None),
    ("AS", "데이터료",              "データ料",                   None),
    ("AT", "★업로드용 팩조건",     "★アップロード用パック条件", None),
    ("AU", "원가",                  "原価",                       None),
    ("AV", "원가계",                "原価計",                     None),
    ("AW", "비고",                  "備考",                       None),
    ("AX", "장합료%",               "帳合料％",                   None),
    ("AY", "본부장 가격",           "本部長価格",                 "本部長価格"),
    ("AZ", "판촉비 소비세",         "販促費消費税",               None),
    ("BA", "품의서 가격",           "稟議書価格",                 None),
    ("BB", "비고2",                 "備考",                       None),
]

# 미리보기용 JP 헤더 목록
PREVIEW_COLUMNS = [jp for _, _, jp, _ in _COLS]


def _cell_val(row: dict, key: str | None):
    if key is None:
        return None
    return row.get(key)


async def _load_rows(doc_ids: list[str]) -> list[dict]:
    settings = get_settings()
    all_rows: list[dict] = []
    for doc_id in doc_ids:
        path: Path = settings.extracted_dir / doc_id / "phase4_output.json"
        if not path.exists():
            raise HTTPException(status_code=404, detail=f"{doc_id}: phase4_output.json 없음")
        data = json.loads(path.read_text(encoding="utf-8"))
        for row in data.get("rows", []):
            all_rows.append({**row, "__doc_id": doc_id})
    return all_rows


# ── 엔드포인트 ────────────────────────────────────────────────────────────────

@router.get("/confirmed-docs")
async def list_confirmed_docs(
    year:  int | None = Query(None),
    month: int | None = Query(None),
    user: dict = Depends(get_current_user),
):
    """확정된 문서 목록 (연월 필터 가능)."""
    from datetime import datetime
    docs = await list_documents()
    result = []
    for d in docs:
        ca = d.get("confirmed_at")
        if not ca:
            continue
        try:
            dt = datetime.fromisoformat(ca)
            if year and dt.year != year:
                continue
            if month and dt.month != month:
                continue
        except (ValueError, TypeError):
            pass
        result.append({
            "doc_id": d["doc_id"],
            "pdf_filename": d.get("pdf_filename", ""),
            "confirmed_at": ca,
            "created_at": d.get("created_at"),
        })
    return result


class DocIdsBody(BaseModel):
    doc_ids: list[str]


@router.post("/preview")
async def preview_rows(body: DocIdsBody, user: dict = Depends(get_current_user)):
    """선택된 문서들의 합산 행 미리보기 (JSON)."""
    if not body.doc_ids:
        return {"columns": PREVIEW_COLUMNS, "rows": []}
    rows = await _load_rows(body.doc_ids)
    preview = [
        {jp: _cell_val(row, key) for _, _, jp, key in _COLS}
        for row in rows
    ]
    return {"columns": PREVIEW_COLUMNS, "rows": preview}


@router.post("/download")
async def download_excel(body: DocIdsBody, user: dict = Depends(get_current_user)):
    """선택된 문서들을 하나의 엑셀 파일로 내보내기."""
    if not body.doc_ids:
        raise HTTPException(status_code=400, detail="문서를 선택하세요")

    rows = await _load_rows(body.doc_ids)

    wb = Workbook()
    ws = wb.active
    ws.title = "入力用"

    fill_ko  = PatternFill("solid", fgColor="0A6E6E")
    fill_jp  = PatternFill("solid", fgColor="E0F0F0")
    font_ko  = Font(bold=True, color="FFFFFF", size=9)
    font_jp  = Font(bold=True, color="0A6E6E", size=9)
    align_c  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_ltr, ko_hdr, jp_hdr, _ in _COLS:
        ci = column_index_from_string(col_ltr)
        c1 = ws.cell(row=1, column=ci, value=ko_hdr)
        c1.fill = fill_ko; c1.font = font_ko; c1.alignment = align_c
        c2 = ws.cell(row=2, column=ci, value=jp_hdr)
        c2.fill = fill_jp; c2.font = font_jp; c2.alignment = align_c

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22

    for r_idx, row in enumerate(rows, start=3):
        for col_ltr, _, _, key in _COLS:
            ci = column_index_from_string(col_ltr)
            ws.cell(row=r_idx, column=ci, value=_cell_val(row, key))

    for col_ltr, _, _, _ in _COLS:
        ws.column_dimensions[col_ltr].width = 11

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    hatsu = rows[0].get("発生月", "") if rows else ""
    suffix = hatsu.replace(".", "") if hatsu else date.today().strftime("%Y%m")
    filename = f"SAP_{suffix}_export.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
