"""
scripts/export_excel.py — Phase 4 결과 → 入力用 Excel 파일 생성
docs/output-format.md 기준. openpyxl 사용.

사용법:
  python3 scripts/export_excel.py --doc "_分東日本_2025.01 (1)"
출력:
  extracted/{doc_id}/{doc_id}_入力用.xlsx
"""
import argparse, json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent.parent

# (col_letter, JP명, KO명, p4_key)
# p4_key = "__xxx" → 특수 계산값, None → 빈 칸
COL_DEFS = [
    ("A",  "",                             "",                    None),
    ("B",  "受注先",                       "판매처 원문명",       "受注先"),
    ("C",  "受注先コード",                 "판매처코드",          "受注先コード"),
    ("D",  "担当者",                       "SP 영업사원",         "担当者"),
    ("E",  "受注先コード（移管含む）",     "판매처코드(이관포함)", None),
    ("F",  "移管受注先",                   "이전판매처",          None),
    ("G",  "移管後担当者",                 "이관후 담당자",       None),
    ("H",  "ロットアウト判別用",           "로트아웃 판별용",     "ロットアウト料別用"),
    ("I",  "PY",                           "PY",                  None),
    ("J",  "代表スーパー",                 "소매처코드",          "代表スーパー"),
    ("K",  "スーパー",                     "소매처명",            "スーパー"),
    ("L",  "商品名",                       "제품명",              "商品名"),
    ("M",  "商品コード",                   "제품코드",            "商品コード"),
    ("N",  "ケース (入数)",                "케이스 입수",         "__case_qty"),
    ("O",  "ボール (入数)",                "볼 입수",             None),
    ("P",  "ケース (数量)",                "케이스 수량",         "ケース"),
    ("Q",  "*",                            "ケース計",            "__q_keesu_kei"),
    ("R",  "ボール (数量)",                "볼 수량",             None),
    ("S",  "*",                            "ボール計",            None),
    ("T",  "バラ",                         "낱개 수량",           "バラ"),
    ("U",  "個数計",                       "총 개수",             "個数計"),
    ("V",  "ケース計",                     "케이스 환산",         "ケース計"),
    ("W",  "発生月",                       "발생월",              "発生月"),
    ("X",  "仕切",                         "시키리(단가)",        "仕切"),
    ("Y",  "仕切売上",                     "시키리 매출",         "仕切売上"),
    ("Z",  "条件1（パック）",              "조건1(팩)",           "条件1（バリュー）（パック）"),
    ("AA", "条件2（パック）",              "조건2(팩)",           "条件2（バリュー）（パック）"),
    ("AB", "条件1（ボール）",              "조건1(볼)",           "条件1（ボール）"),
    ("AC", "条件2（ボール）",              "조건2(볼)",           "条件2（ボール）"),
    ("AD", "条件1（ケース）",              "조건1(케이스)",       "条件1（ケース）"),
    ("AE", "条件2（ケース）",              "조건2(케이스)",       "条件2（ケース）"),
    ("AF", "パック条件1",                  "팩조건1(단위통합)",   None),
    ("AG", "パック条件2",                  "팩조건2(단위통합)",   None),
    ("AH", "NET",                          "NET",                 "NET"),
    ("AI", "パック条件1合計",              "팩조건1 합계",        None),
    ("AJ", "パック条件2合計",              "팩조건2 합계",        None),
    ("AK", "差額",                         "차액",                None),
    ("AL", "未収金額合計",                 "지급해야 될 금액",    "__kin_gaku"),
    ("AM", "想定納価",                     "상정 납가",           None),
    ("AN", "ﾌｨｰ％(×納価）①",           "피%(×납가)①",       None),
    ("AO", "ﾌｨｰ％(×納価）②",           "피%(×납가)②",       None),
    ("AP", "最終NET",                      "최종 NET",            None),
    ("AQ", "最終NET売上",                  "최종 NET 매출",       None),
    ("AR", "大陳料",                       "대진료",              None),
    ("AS", "データ料",                     "데이터료",            None),
    ("AT", "★アップロード用パック条件",   "★업로드용 팩조건",   None),
    ("AU", "原価",                         "원가",                None),
    ("AV", "原価計",                       "원가계",              None),
    ("AW", "備考",                         "비고",                None),
    ("AX", "帳合料％",                     "장합료%",             None),
    ("AY", "本部長価格",                   "본부장 가격",         "本部長"),
    ("AZ", "販促費消費税",                 "판촉비 소비세",       None),
    ("BA", "稟議書価格",                   "품의서 가격",         None),
    ("BB", "備考",                         "비고2",               None),
]

# 각 컬럼의 기본 너비
COL_WIDTHS = {
    1: 2,   # A spacer
    2: 22,  # B 受注先
    3: 12,  # C 受注先コード
    4: 10,  # D 担当者
    8: 6,   # H ロットアウト
    10: 12, # J 代表スーパー
    11: 20, # K スーパー
    12: 24, # L 商品名
    13: 13, # M 商品コード
    14: 7,  # N ケース入数
    20: 10, # T バラ
    21: 10, # U 個数計
    22: 10, # V ケース計
    23: 9,  # W 発生月
    24: 8,  # X 仕切
    25: 10, # Y 仕切売上
    34: 10, # AH NET
    38: 14, # AL 未収金額合計
    51: 10, # AY 本部長価格
}


def run(doc_id: str) -> Path:
    doc_dir = BASE / "extracted" / doc_id
    p4_path = doc_dir / "phase4_output.json"
    p3_path = doc_dir / "phase3_output.json"

    for p in (p4_path, p3_path):
        if not p.exists():
            raise FileNotFoundError(f"{p} 없음")

    p4 = json.loads(p4_path.read_text(encoding="utf-8"))
    p3 = json.loads(p3_path.read_text(encoding="utf-8"))
    rows_p4  = p4["rows"]
    items_p3 = p3["items"]

    # ── workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "入力用"

    fill_hdr_ko = PatternFill("solid", fgColor="4472C4")
    fill_hdr_jp = PatternFill("solid", fgColor="2F5496")
    font_hdr    = Font(bold=True, color="FFFFFF", size=9)
    fill_warn   = PatternFill("solid", fgColor="FFD7D7")   # NET < 本部長
    fill_uncfm  = PatternFill("solid", fgColor="FFF2CC")   # unconfirmed
    align_ctr   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_mid   = Alignment(vertical="center")

    # ── 헤더 2행 ──────────────────────────────────────────────────────────────
    for ci, (_, jp, ko, _key) in enumerate(COL_DEFS, 1):
        c1 = ws.cell(row=1, column=ci, value=ko or None)
        c2 = ws.cell(row=2, column=ci, value=jp or None)
        for c, fill in ((c1, fill_hdr_ko), (c2, fill_hdr_jp)):
            c.fill  = fill
            c.font  = font_hdr
            c.alignment = align_ctr

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "B3"

    # ── 데이터 행 ─────────────────────────────────────────────────────────────
    for ri, (r4, i3) in enumerate(zip(rows_p4, items_p3), 3):
        cols3    = i3.get("columns", {})
        case_qty = cols3.get("ケース入数") or 0
        kin_gaku = cols3.get("金額") or 0
        keesu    = r4.get("ケース") or 0
        q_val    = int(keesu * case_qty) if case_qty else 0

        aug = {
            "__case_qty":    case_qty or None,
            "__q_keesu_kei": q_val or None,
            "__kin_gaku":    kin_gaku or None,
        }

        row_fill = (fill_uncfm if r4.get("unconfirmed")
                    else fill_warn if r4.get("net_lt_honbu")
                    else None)

        for ci, (_, _jp, _ko, key) in enumerate(COL_DEFS, 1):
            if key is None:
                v = None
            elif key.startswith("__"):
                v = aug.get(key)
            else:
                v = r4.get(key)

            cell = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = align_mid
            if row_fill:
                cell.fill = row_fill

    # ── 컬럼 너비 ────────────────────────────────────────────────────────────
    for ci in range(1, len(COL_DEFS) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(ci, 9)

    # ── 저장 ─────────────────────────────────────────────────────────────────
    out_path = doc_dir / f"{doc_id}_入力用.xlsx"
    wb.save(out_path)
    print(f"✅ 저장 완료: {out_path}")
    print(f"   데이터 행: {len(rows_p4)}건 / 컬럼: {len(COL_DEFS)}열 (A~BB)")
    return out_path


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Phase 4 → 入力用 Excel 생성")
    ap.add_argument("--doc", required=True, help="doc_id")
    args = ap.parse_args()
    run(args.doc)
